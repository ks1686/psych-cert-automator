"""Ineligibility report generator.

Writes a formatted .xlsx file listing every person whose CE request
could not be fulfilled, with diagnostic details.

Public API:
    generate_ineligibility_report(entries, output_path) → str
"""

from __future__ import annotations

from collections import Counter
from pathlib import Path
from typing import TYPE_CHECKING, assert_never

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from src.models.certificate import EligibilityStatus

if TYPE_CHECKING:
    from openpyxl.worksheet.worksheet import Worksheet

    from src.models.certificate import IneligibilityEntry

# ═══════════════════════════════════════════════════════════════════════════════
# Column definitions
# ═══════════════════════════════════════════════════════════════════════════════

_NAME_QUALT_COL = "Name (Qualtrics)"
_NAME_ZOOM_COL = "Name (Zoom)"
_MATCH_STATUS_COL = "Match Status"
_LATE_JOIN_COL = "Late Join (min)"
_EARLY_LEAVE_COL = "Early Leave (min)"
_GAPS_COL = "Mid-Session Gaps (min)"
_TOTAL_MISSED_COL = "Total Missed (min)"
_REJECTED_CE_COL = "Rejected CE Types"
_STATUS_COL = "Status"
_REASON_COL = "Reason"

_HEADERS: tuple[str, ...] = (
    _NAME_QUALT_COL,
    _NAME_ZOOM_COL,
    _MATCH_STATUS_COL,
    _LATE_JOIN_COL,
    _EARLY_LEAVE_COL,
    _GAPS_COL,
    _TOTAL_MISSED_COL,
    _REJECTED_CE_COL,
    _STATUS_COL,
    _REASON_COL,
)

_NUMBER_COLS: frozenset[str] = frozenset(
    {_LATE_JOIN_COL, _EARLY_LEAVE_COL, _GAPS_COL, _TOTAL_MISSED_COL}
)

# ═══════════════════════════════════════════════════════════════════════════════
# Style constants
# ═══════════════════════════════════════════════════════════════════════════════

_HEADER_FONT = Font(bold=True, size=11)
_HEADER_FILL = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
_HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)

_TEXT_ALIGNMENT = Alignment(horizontal="left", vertical="center")
_NUMBER_ALIGNMENT = Alignment(horizontal="right", vertical="center")

_BORDER_BOTTOM = Border(bottom=Side(style="thin"))

_STRIPE_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

# Status color fills
_FILL_NOT_FOUND = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
_FILL_AMBIGUOUS = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
_FILL_INSUFFICIENT = PatternFill(start_color="F4B942", end_color="F4B942", fill_type="solid")

# Sheet names
_INELIGIBLE_SHEET = "Ineligible"
_SUMMARY_SHEET = "Summary"
_FORMULA_PREFIXES = ("=", "+", "-", "@")


# ═══════════════════════════════════════════════════════════════════════════════
# Private helpers
# ═══════════════════════════════════════════════════════════════════════════════


def _status_fill(status: EligibilityStatus) -> PatternFill:
    """Return the row fill colour for a given eligibility status.

    Uses ``match/case`` with ``assert_never`` so the type checker enforces
    exhaustive coverage of all ``EligibilityStatus`` variants.
    """
    match status:
        case EligibilityStatus.NOT_FOUND_IN_ATTENDANCE:
            return _FILL_NOT_FOUND
        case EligibilityStatus.NAME_MATCH_AMBIGUOUS:
            return _FILL_AMBIGUOUS
        case EligibilityStatus.ATTENDANCE_INSUFFICIENT:
            return _FILL_INSUFFICIENT
        case EligibilityStatus.ELIGIBLE:
            return PatternFill()
        case _:
            assert_never(status)


def _blank(value: int | None) -> int:
    """Return 0 for None, otherwise the value."""
    return value if value is not None else 0


def _xlsx_text(value: str) -> str:
    if value.startswith(_FORMULA_PREFIXES):
        return f"'{value}"
    return value


def _build_data_row(entry: IneligibilityEntry) -> list[str | int]:
    """Build a single data row for the ineligibility sheet.

    Args:
        entry: A single ``IneligibilityEntry`` to render.

    Returns:
        A list of 10 cell values matching the column order of ``_HEADERS``.
    """
    total_missed = (
        _blank(entry.late_join_minutes)
        + _blank(entry.early_leave_minutes)
        + _blank(entry.total_gaps_minutes)
    )
    return [
        _xlsx_text(entry.name_qualtrics),
        _xlsx_text(entry.name_zoom or ""),
        _xlsx_text(entry.match_status),
        entry.late_join_minutes if entry.late_join_minutes is not None else "",
        entry.early_leave_minutes if entry.early_leave_minutes is not None else "",
        entry.total_gaps_minutes if entry.total_gaps_minutes is not None else "",
        total_missed if total_missed > 0 else "",
        _xlsx_text(", ".join(entry.rejected_ce_types)),
        _xlsx_text(entry.status.value),
        _xlsx_text(entry.reason),
    ]


def _apply_data_row_style(
    ws: Worksheet,
    row_num: int,
    entry: IneligibilityEntry,
    col_to_idx: dict[str, int],
) -> None:
    """Apply per-cell alignment, fill, and stripe formatting to a data row.

    The Status column receives colour-coding by ``EligibilityStatus``.
    Other columns alternate white / light-gray for readability.
    Number columns are right-aligned; text columns are left-aligned.
    """
    status_fill = _status_fill(entry.status)
    is_striped = row_num % 2 == 0

    for col_name, col_idx in col_to_idx.items():
        cell = ws.cell(row=row_num, column=col_idx)
        cell.alignment = (
            _NUMBER_ALIGNMENT if col_name in _NUMBER_COLS else _TEXT_ALIGNMENT
        )
        if col_name == _STATUS_COL:
            cell.fill = status_fill
        elif is_striped:
            cell.fill = _STRIPE_FILL


def _auto_fit_columns(ws: Worksheet) -> None:
    """Set column widths to fit the longest value in each column.

    Clamped to a maximum of 50 characters to prevent obscenely wide columns.
    """
    for col_cells in ws.columns:
        max_length = 0
        col_idx = col_cells[0].column
        if col_idx is None:
            continue
        col_letter = get_column_letter(col_idx)
        for cell in col_cells:
            if cell.value is not None:
                cell_len = len(str(cell.value))
                max_length = max(max_length, cell_len)
        adjusted = min(max_length + 2, 50)
        ws.column_dimensions[col_letter].width = adjusted


def _write_summary_sheet(
    ws: Worksheet, entries: list[IneligibilityEntry]
) -> None:
    """Populate the Summary sheet with aggregate statistics.

    Includes:
    - Total ineligible request count
    - Breakdown by ``EligibilityStatus`` (count per variant)
    - Breakdown by rejected CE type (count per type)
    """
    header_font = Font(bold=True, size=12)
    section_font = Font(bold=True, size=11)

    # ── Overall counts ──
    _ = ws.cell(row=1, column=1, value="Ineligibility Report Summary")
    ws["A1"].font = header_font
    _ = ws.cell(row=3, column=1, value="Total Ineligible Requests")
    ws["A3"].font = section_font
    _ = ws.cell(row=3, column=2, value=len(entries))

    # ── Breakdown by EligibilityStatus ──
    _ = ws.cell(row=5, column=1, value="Breakdown by Status")
    ws["A5"].font = header_font
    _ = ws.cell(row=6, column=1, value="Status")
    ws["A6"].font = section_font
    _ = ws.cell(row=6, column=2, value="Count")
    ws["B6"].font = section_font

    status_counter: dict[str, int] = dict(
        Counter(entry.status.value for entry in entries)
    )
    row = 7
    for status_value, count in sorted(status_counter.items()):
        _ = ws.cell(row=row, column=1, value=status_value)
        _ = ws.cell(row=row, column=2, value=count)
        row += 1

    # ── Breakdown by CE Type ──
    row += 1
    _ = ws.cell(row=row, column=1, value="Breakdown by CE Type Rejected")
    ws.cell(row=row, column=1).font = header_font
    row += 1
    _ = ws.cell(row=row, column=1, value="CE Type")
    ws.cell(row=row, column=1).font = section_font
    _ = ws.cell(row=row, column=2, value="Count")
    ws.cell(row=row, column=2).font = section_font

    ce_counter: Counter[str] = Counter()
    for entry in entries:
        for ce_type in entry.rejected_ce_types:
            ce_counter[ce_type] += 1

    row += 1
    for ce_type, count in sorted(ce_counter.items()):
        _ = ws.cell(row=row, column=1, value=ce_type)
        _ = ws.cell(row=row, column=2, value=count)
        row += 1

    # Fixed column widths for the summary sheet
    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 15


# ═══════════════════════════════════════════════════════════════════════════════
# Public API
# ═══════════════════════════════════════════════════════════════════════════════


def generate_ineligibility_report(
    entries: list[IneligibilityEntry],
    output_path: str,
) -> str:
    """Generate an ineligibility summary ``.xlsx`` report.

    Writes a formatted workbook with two sheets:

    - **Ineligible** — one row per entry with name, match status, attendance
      metrics, requested CE types, status, and reason. Includes bold headers,
      auto-filter, frozen header row, auto-fit columns, and status-colour-coded
      rows.

    - **Summary** — aggregate counts: total ineligible, breakdown by
      ``EligibilityStatus``, and breakdown by rejected CE type.

    Args:
        entries: List of ``IneligibilityEntry`` objects to include.
        output_path: File path where the ``.xlsx`` file is written. Parent
            directories are created if they do not exist.

    Returns:
        The *output_path* as a string.

    Raises:
        FileNotFoundError: If a parent directory cannot be created and does
            not exist (e.g. permission denied on a read-only filesystem).
    """
    col_to_idx = {header: idx + 1 for idx, header in enumerate(_HEADERS)}

    wb = Workbook()
    ws = wb.active
    if ws is None:
        _no_sheet_msg = "Workbook has no active worksheet"
        raise RuntimeError(_no_sheet_msg)
    ws.title = _INELIGIBLE_SHEET

    # ── Header row ──
    for col_name, col_idx in col_to_idx.items():
        _ = ws.cell(row=1, column=col_idx, value=col_name)
        ws.cell(row=1, column=col_idx).font = _HEADER_FONT
        ws.cell(row=1, column=col_idx).fill = _HEADER_FILL
        ws.cell(row=1, column=col_idx).alignment = _HEADER_ALIGNMENT
        ws.cell(row=1, column=col_idx).border = _BORDER_BOTTOM

    # ── Data rows ──
    for i, entry in enumerate(entries):
        row_num = i + 2  # 1-indexed, header occupies row 1
        row_data = _build_data_row(entry)
        for col_idx in col_to_idx.values():
            _ = ws.cell(row=row_num, column=col_idx, value=row_data[col_idx - 1])
        _apply_data_row_style(ws, row_num, entry, col_to_idx)

    # ── Finalize ineligible sheet ──
    _auto_fit_columns(ws)
    ws.freeze_panes = "A2"
    last_data_row = len(entries) + 1
    ws.auto_filter.ref = f"A1:{get_column_letter(len(_HEADERS))}{last_data_row}"

    # ── Summary sheet ──
    wb.create_sheet(title=_SUMMARY_SHEET)
    _write_summary_sheet(wb[_SUMMARY_SHEET], entries)

    # ── Save ──
    out = Path(output_path)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(out))

    return str(out)
