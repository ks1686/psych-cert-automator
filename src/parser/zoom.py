"""Zoom attendance report parser — reads .xlsx into typed participant data.

Parses a Zoom "Usage Report" attendance .xlsx export, extracts session
metadata, filters out waiting-room segments, groups records by normalized
participant name, and returns one ``ParticipantAttendance`` per unique person.
"""

from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime
from typing import override

import openpyxl

from src.models.participant import (
    AttendanceRecord,
    ParticipantAttendance,
    ParticipantDataError,
    normalize_name,
)


@dataclass(frozen=True, slots=True)
class ZoomSession:
    """Container for a parsed Zoom attendance report.

    Attributes:
        session_start: When the Zoom meeting started.
        session_end: When the Zoom meeting ended.
        participants: One ``ParticipantAttendance`` per unique person,
            sorted by ``first_join`` ascending.
    """

    session_start: datetime
    session_end: datetime
    participants: tuple[ParticipantAttendance, ...]


@dataclass(frozen=True, slots=True)
class ZoomParseError(Exception):
    """Error raised when a Zoom attendance report cannot be parsed.

    Carries the filepath and a human-readable reason so callers can surface
    precise diagnostics.
    """

    filepath: str
    reason: str

    @override
    def __str__(self) -> str:
        """Return a human-readable error message."""
        return f"Zoom parse error for {self.filepath!r}: {self.reason}"


def parse_zoom_attendance(filepath: str) -> ZoomSession:
    """Parse a Zoom attendance .xlsx report into typed participant data.

    Opens the workbook, extracts session start/end times from the metadata
    rows, reads per-segment participant records, filters out waiting-room
    segments, groups records by normalized participant name, and returns
    aggregated ``ParticipantAttendance`` objects.

    Args:
        filepath: Path to the Zoom-generated ``.xlsx`` attendance report.

    Returns:
        A ``ZoomSession`` containing session times and one
        ``ParticipantAttendance`` per unique person.

    Raises:
        ZoomParseError: If the file is missing, not a valid .xlsx, or
            contains unparseable data.
    """
    workbook = _open_workbook(filepath)

    ws = workbook.active
    if ws is None:
        raise ZoomParseError(filepath=filepath, reason="workbook has no active sheet")

    raw_start: object = ws["E2"].value  # pyright: ignore[reportAny]  # openpyxl stubs
    raw_end: object = ws["F2"].value  # pyright: ignore[reportAny]  # openpyxl stubs
    session_start = _parse_datetime_cell(raw_start, "session_start", filepath)
    session_end = _parse_datetime_cell(raw_end, "session_end", filepath)

    raw_records: list[AttendanceRecord] = []
    for row in ws.iter_rows(min_row=5, values_only=True):
        if row[0] is None:
            continue
        raw_records.append(_parse_attendance_row(row, filepath))

    workbook.close()

    # Filter out waiting-room segments, then group by normalized name.
    active = [r for r in raw_records if not r.is_waiting_room]

    by_name: dict[str, list[AttendanceRecord]] = {}
    for record in active:
        key = normalize_name(record.name_raw)
        if key in by_name:
            by_name[key].append(record)
        else:
            by_name[key] = [record]

    participants: list[ParticipantAttendance] = []
    for group in by_name.values():
        try:
            participants.append(ParticipantAttendance.from_records(group))
        except ParticipantDataError:
            # Should not happen after filtering, but skip gracefully.
            continue

    participants.sort(key=lambda p: p.first_join)

    return ZoomSession(
        session_start=session_start,
        session_end=session_end,
        participants=tuple(participants),
    )


# ── Internal helpers ──────────────────────────────────────────────────────────


def _open_workbook(filepath: str) -> openpyxl.Workbook:
    """Open an .xlsx workbook with error handling.

    Args:
        filepath: Path to the .xlsx file.

    Returns:
        An open ``openpyxl.Workbook``. The caller is responsible for closing it.

    Raises:
        ZoomParseError: If the file is missing or not a valid .xlsx.
    """
    try:
        return openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    except FileNotFoundError:
        raise ZoomParseError(filepath=filepath, reason="file not found") from None
    except KeyError as exc:
        raise ZoomParseError(
            filepath=filepath, reason=f"unexpected workbook structure: {exc}"
        ) from exc


def _parse_datetime_cell(
    value: object,
    label: str,
    filepath: str,
) -> datetime:
    """Parse a cell value that must be a ``datetime``.

    Args:
        value: Raw cell value from openpyxl.
        label: Human-readable field name for error messages.
        filepath: Source file for error context.

    Returns:
        The parsed ``datetime``.

    Raises:
        ZoomParseError: If the value is not a ``datetime``.
    """
    if isinstance(value, datetime):
        return value
    raise ZoomParseError(
        filepath=filepath,
        reason=f"expected datetime for {label}, got {type(value).__name__}: {value!r}",
    )


def _parse_attendance_row(
    row: tuple[object, ...],
    filepath: str,
) -> AttendanceRecord:
    """Parse a single participant data row into an ``AttendanceRecord``.

    Expects a 7-element tuple in Zoom report column order: name, email,
    join_time, leave_time, duration, guest, waiting_room.

    Args:
        row: Raw row tuple from openpyxl ``iter_rows(values_only=True)``.
        filepath: Source file for error context.

    Returns:
        A validated ``AttendanceRecord``.

    Raises:
        ZoomParseError: If any cell contains unparseable data.
    """
    name_raw = _parse_name_cell(row[0], filepath)
    email = _parse_email_cell(row[1])
    join_time = _parse_datetime_cell(row[2], f"join_time for {name_raw!r}", filepath)
    leave_time = _parse_datetime_cell(row[3], f"leave_time for {name_raw!r}", filepath)
    duration = _parse_duration_cell(row[4], name_raw, filepath)
    is_guest = _parse_yes_no(row[5], "guest", name_raw, filepath)
    is_waiting_room = _parse_yes_no(row[6], "waiting room", name_raw, filepath)

    return AttendanceRecord(
        name_raw=name_raw,
        email=email,
        join_time=join_time,
        leave_time=leave_time,
        duration_minutes=duration,
        is_guest=is_guest,
        is_waiting_room=is_waiting_room,
    )


def _parse_name_cell(value: object, filepath: str) -> str:
    """Parse a name cell value — must be a non-empty string.

    Args:
        value: Raw cell value.
        filepath: Source file for error context.

    Returns:
        The name string, with leading/trailing whitespace stripped.

    Raises:
        ZoomParseError: If the value is not a non-empty string.
    """
    if isinstance(value, str) and value.strip():
        return value.strip()
    raise ZoomParseError(
        filepath=filepath,
        reason=f"expected non-empty string for name, got {value!r}",
    )


def _parse_email_cell(value: object) -> str | None:
    """Parse an email cell value — string or None.

    Args:
        value: Raw cell value (may be ``str``, ``None``, or other).

    Returns:
        The email string, or ``None`` if the cell was empty or non-string.
    """
    return value if isinstance(value, str) else None


def _parse_duration_cell(
    value: object,
    name: str,
    filepath: str,
) -> int:
    """Parse a duration cell value — must be a non-negative integer.

    Args:
        value: Raw cell value (typically ``int`` from Zoom).
        name: Participant name for error context.
        filepath: Source file for error context.

    Returns:
        Duration in minutes as ``int``.

    Raises:
        ZoomParseError: If the value is not a non-negative number.
    """
    match value:
        case int() | float():
            if value >= 0:
                return int(value)
            raise ZoomParseError(
                filepath=filepath,
                reason=f"negative duration for {name!r}: {value}",
            )
        case _:
            raise ZoomParseError(
                filepath=filepath,
                reason=(
                    f"expected number for duration of {name!r}, "
                    f"got {type(value).__name__}: {value!r}"
                ),
            )


def _parse_yes_no(
    value: object,
    field: str,
    name: str,
    filepath: str,
) -> bool:
    """Parse a 'Yes'/'No' string cell into a boolean.

    Args:
        value: Raw cell value (expected ``"Yes"`` or ``"No"``).
        field: Field label for error messages.
        name: Participant name for error context.
        filepath: Source file for error context.

    Returns:
        ``True`` for ``"Yes"``, ``False`` for ``"No"``.

    Raises:
        ZoomParseError: If the value is not ``"Yes"`` or ``"No"``.
    """
    match value:
        case "Yes":
            return True
        case "No":
            return False
        case _:
            raise ZoomParseError(
                filepath=filepath,
                reason=(
                    f"expected 'Yes' or 'No' for {field} of {name!r}, "
                    f"got {value!r}"
                ),
            )
