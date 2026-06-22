"""Parser for Qualtrics CE request .xlsx exports.

Reads a Qualtrics survey export and produces typed ``CERequest`` objects,
auto-detecting column structure via header keyword matching. Supports
single-select, multi-select (comma-separated), and checkbox-style CE
type columns.
"""

from __future__ import annotations

import re
from dataclasses import dataclass
from enum import StrEnum, auto, unique
from pathlib import Path
from typing import TYPE_CHECKING

if TYPE_CHECKING:
    from collections.abc import Iterator

from openpyxl import load_workbook

from src.models.certificate import CERequest
from src.models.training import CEType

# ── Column detection constants ─────────────────────────────────────────────

_CE_QUESTION_KW: tuple[str, ...] = ("type of ce credit",)
_CE_SELECTED_KW: tuple[str, ...] = ("selected choice",)
_LICENSE_KW: tuple[str, ...] = (
    "license #", "license number", "certificate #", "certificate number",
)
_NAME_KW: tuple[str, ...] = ("name (as you would like", "name on certificate")
_EMAIL_KW: tuple[str, ...] = ("preferred email",)
_LICENSE_PROMPT_RE: re.Pattern[str] = re.compile(
    r"\([^)]*(?:license|certificate|enter|#)[^)]*\)", re.IGNORECASE,
)
"""Match a parenthetical mentioning a license or certificate."""


@unique
class _CeColKind(StrEnum):
    """Classification of a CE-related survey column."""

    SELECTION = auto()  # Main dropdown or per-type checkbox selection
    LICENSE = auto()    # Text entry for license/certificate number
    UNKNOWN = auto()    # CE-related but unrecognised format


# ── Header parsing helpers ─────────────────────────────────────────────────


def _contains(header: str, keywords: tuple[str, ...]) -> bool:
    """Check whether any *keyword* appears (case-insensitive) in *header*."""
    lower = header.lower()
    return any(kw in lower for kw in keywords)


def _extract_ce_name(header: str, kind: _CeColKind) -> str | None:
    """Extract CE type name from a license or checkbox column header.

    Header format: ``<question text> - <CE Type Name> (<prompt>) - Text``.

    The CE type name may itself contain parentheses (e.g.
    ``"Psychologist (New York)"``). For license columns we locate the
    trailing prompt parenthetical and return text between the preceding
    `` - `` separator and that prompt. For selection columns the segment
    after the last `` - `` *is* the CE type name.
    """
    text = header.strip()
    if text.endswith(" - Text"):
        text = text.removesuffix(" - Text")

    sep = text.rfind(" - ")
    if sep == -1:
        return None
    segment = text[sep + 3 :].strip()

    # For selection columns the segment may be the plain dropdown
    # (``"Selected Choice"``) or a checkbox column
    # (``"Psychologist (APA) - Selected Choice"``).  Extract the CE
    # name from the checkbox form, return None for the plain dropdown.
    if kind is _CeColKind.SELECTION:
        if segment.lower() == "selected choice":
            return None  # plain master dropdown — no CE name in header
        suffix = " - selected choice"
        if segment.lower().endswith(suffix):
            segment = segment[: -len(suffix)].strip()
        return segment or None

    # For license columns: find prompt parenthetical, return text before it.
    match = _LICENSE_PROMPT_RE.search(segment)
    if match:
        return segment[: match.start()].strip() or None

    return None


def _classify_ce_column(header: str) -> _CeColKind:
    """Classify a CE-question column as selection, license, or unknown."""
    if _contains(header, _CE_SELECTED_KW):
        return _CeColKind.SELECTION
    if _contains(header, _LICENSE_KW) or _LICENSE_PROMPT_RE.search(header):
        return _CeColKind.LICENSE
    return _CeColKind.UNKNOWN


# ── Column detection ───────────────────────────────────────────────────────


@dataclass(frozen=True, slots=True)
class _ColumnLayout:
    """Detected column indices for a Qualtrics export."""

    name_col: int
    email_col: int
    # Main CE selection column (single/multi-select format). None in checkbox format.
    main_ce_col: int | None
    # Mapping: CE type display name → license number column index.
    license_cols: dict[str, int]
    # In checkbox format: list of (ce_type_name, column_index).
    ce_checkbox_cols: list[tuple[str, int]]


def _detect_columns(headers: list[str]) -> _ColumnLayout:
    """Scan header row and return a ``_ColumnLayout`` of detected column indices.

    Raises:
        ValueError: If required columns (name, email) cannot be detected.
    """
    name_col: int | None = None
    email_col: int | None = None
    license_cols: dict[str, int] = {}
    ce_checkbox_cols: list[tuple[str, int]] = []

    for i, raw in enumerate(headers):
        hdr = raw or ""

        if name_col is None and _contains(hdr, _NAME_KW):
            name_col = i
            continue
        if email_col is None and _contains(hdr, _EMAIL_KW):
            email_col = i
            continue

        _scan_ce_column(hdr, i, license_cols, ce_checkbox_cols)

    main_ce_col: int | None = (
        None if ce_checkbox_cols else _find_plain_selection(headers)
    )

    if name_col is None:
        msg = f"Missing name column (keywords: {_NAME_KW})"
        raise ValueError(msg)
    if email_col is None:
        msg = f"Missing email column (keywords: {_EMAIL_KW})"
        raise ValueError(msg)

    return _ColumnLayout(
        name_col=name_col,
        email_col=email_col,
        main_ce_col=main_ce_col,
        license_cols=license_cols,
        ce_checkbox_cols=ce_checkbox_cols,
    )


def _scan_ce_column(
    header: str,
    col_idx: int,
    license_cols: dict[str, int],
    ce_checkbox_cols: list[tuple[str, int]],
) -> None:
    """If *header* is a CE-question column, classify it and record its index."""
    if not _contains(header, _CE_QUESTION_KW):
        return
    kind = _classify_ce_column(header)
    ce_name = _extract_ce_name(header, kind)
    if ce_name is None:
        return
    match kind:
        case _CeColKind.SELECTION:
            ce_checkbox_cols.append((ce_name, col_idx))
        case _CeColKind.LICENSE:
            license_cols[ce_name] = col_idx
        case _CeColKind.UNKNOWN:
            pass


def _find_plain_selection(headers: list[str]) -> int | None:
    """Find a plain ``Selected Choice`` column (no CE type in its header)."""
    for i, raw in enumerate(headers):
        hdr = (raw or "").lower()
        if all(kw in hdr for kw in (*_CE_QUESTION_KW, *_CE_SELECTED_KW)):
            return i
    return None


# ── Cell normalization ─────────────────────────────────────────────────────


def _cell_str(value: object) -> str | None:
    """Normalize an openpyxl cell value to a stripped string, or ``None``."""
    if value is None:
        return None
    return str(value).strip() or None


def _split_ce_types(raw: str) -> list[str]:
    """Split comma-separated CE type selection into individual types."""
    return [t.strip() for t in raw.split(",") if t.strip()]


# ── Row parsing ────────────────────────────────────────────────────────────


def _parse_row(
    row_values: list[object],
    layout: _ColumnLayout,
) -> Iterator[CERequest]:
    """Yield one ``CERequest`` per CE type found in a single survey response row."""
    name = _cell_str(row_values[layout.name_col])
    if not name:
        return

    email = _cell_str(row_values[layout.email_col])

    match layout:
        case _ColumnLayout(main_ce_col=int(ce_col)) if ce_col >= 0:
            raw_ce = _cell_str(row_values[ce_col])
            if raw_ce:
                yield from _from_selection(name, email, raw_ce, layout)
        case _ColumnLayout(ce_checkbox_cols=cols) if cols:
            yield from _from_checkboxes(name, email, row_values, cols, layout)
        case _:
            return


def _from_selection(
    name: str,
    email: str | None,
    raw_ce: str,
    layout: _ColumnLayout,
) -> Iterator[CERequest]:
    """Yield CERequests from a single/multi-select CE dropdown value."""
    for ce_name in _split_ce_types(raw_ce):
        license_number = _lookup_license(ce_name, layout.license_cols, None)
        yield CERequest(
            name_on_certificate=name,
            email=email,
            ce_type=CEType(ce_name),
            license_number=license_number,
        )


def _from_checkboxes(
    name: str,
    email: str | None,
    row_values: list[object],
    cols: list[tuple[str, int]],
    layout: _ColumnLayout,
) -> Iterator[CERequest]:
    """Yield CERequests from checkbox-style CE type columns."""
    for ce_name, col_idx in cols:
        value = _cell_str(row_values[col_idx])
        if not value:
            continue
        license_number = _lookup_license(ce_name, layout.license_cols, row_values)
        yield CERequest(
            name_on_certificate=name,
            email=email,
            ce_type=CEType(ce_name),
            license_number=license_number,
        )


def _lookup_license(
    ce_name: str,
    license_cols: dict[str, int],
    row_values: list[object] | None,
) -> str | None:
    """Return the license number for *ce_name*, or ``None``."""
    col_idx = license_cols.get(ce_name)
    if col_idx is None or row_values is None:
        return None
    return _cell_str(row_values[col_idx])


# ── Public API ──────────────────────────────────────────────────────────────


def parse_qualtrics_export(filepath: str) -> list[CERequest]:
    """Parse a Qualtrics CE request .xlsx export into typed ``CERequest`` objects.

    Auto-detects column structure by scanning header row keywords for name,
    email, CE type selection (single/multi-select/checkbox), and license
    number columns.

    Multi-select CE responses (comma-separated) are split into individual
    ``CERequest`` entries, one per CE type per person.

    Args:
        filepath: Path to a ``.xlsx`` file exported from Qualtrics.

    Returns:
        One ``CERequest`` per person per CE type.

    Raises:
        FileNotFoundError: If *filepath* does not exist.
        ValueError: If required columns cannot be detected.
    """
    path = Path(filepath)
    if not path.exists():
        msg = f"Qualtrics export not found: {filepath}"
        raise FileNotFoundError(msg)

    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb.active
        if ws is None:
            msg = f"No active worksheet in {filepath}"
            raise ValueError(msg)

        headers: list[str] = [
            str(cell.value) if cell.value is not None else ""
            for cell in next(ws.iter_rows(min_row=1, max_row=1))
        ]
        layout = _detect_columns(headers)

        requests: list[CERequest] = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            row_values: list[object] = list(row)
            requests.extend(_parse_row(row_values, layout))

        return requests
    finally:
        wb.close()
